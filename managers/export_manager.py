import csv
import os
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
import logging
from managers.data_manager import DataManager

# ExportManager class content

class ExportManager:
    """Manages data export to various formats"""

    def __init__(self, data_manager: DataManager):
        self.data_manager = data_manager
        self.export_log: List[Dict] = []

    def log_export(self, format_type: str, filename: str, success: bool) -> None:
        """Log export operation"""
        log_entry = {
            'timestamp': datetime.now().isoformat(),
            'format': format_type,
            'filename': filename,
            'success': success
        }
        self.export_log.append(log_entry)
        logging.info(f"Export {format_type} to {filename}: {'Success' if success else 'Failed'}")

    def export_to_xlsx(self, filename: str = "eduplatform_data.xlsx") -> bool:
        """Export all data to the Excel file"""
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Users sheet
            users_sheet = wb.create_sheet("Users")
            users_sheet.append(["ID", "Full Name", "Email", "Role", "Created At", "Phone", "Address"])

            for user in self.data_manager.users.values():
                profile = user.get_profile()
                users_sheet.append([
                    profile['id'], profile['full_name'], profile['email'],
                    profile['role'], profile['created_at'], profile['phone'], profile['address']
                ])

            # Students sheet
            students_sheet = wb.create_sheet("Students")
            students_sheet.append(["User ID", "Grade", "Subjects", "Average Grade"])

            for student in self.data_manager.students.values():
                avg_grade = student.calculate_average_grade()
                students_sheet.append([
                    student._id, student.grade,
                    ", ".join(student.subjects.keys()), avg_grade
                ])

            # Teachers sheet
            teachers_sheet = wb.create_sheet("Teachers")
            teachers_sheet.append(["User ID", "Subjects", "Classes", "Workload"])

            for teacher in self.data_manager.teachers.values():
                teachers_sheet.append([
                    teacher._id, ", ".join(teacher.subjects),
                    ", ".join(teacher.classes), teacher.workload
                ])

            # Assignments sheet
            assignments_sheet = wb.create_sheet("Assignments")
            assignments_sheet.append([
                "ID", "Title", "Subject", "Teacher ID", "Class ID",
                "Deadline", "Difficulty", "Submissions", "Grades"
            ])

            for assignment in self.data_manager.assignments.values():
                assignments_sheet.append([
                    assignment.id, assignment.title, assignment.subject,
                    assignment.teacher_id, assignment.class_id, assignment.deadline,
                    assignment.difficulty, len(assignment.submissions), len(assignment.grades)
                ])

            wb.save(filename)
            self.log_export("XLSX", filename, True)
            return True

        except Exception as e:
            logging.error(f"XLSX export failed: {e}")
            self.log_export("XLSX", filename, False)
            return False

    def export_to_csv(self, directory: str = "csv_exports") -> bool:
        """Export data to CSV files"""
        try:
            os.makedirs(directory, exist_ok=True)

            # Users CSV
            users_file = os.path.join(directory, "users.csv")
            with open(users_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "Full Name", "Email", "Role", "Created At", "Phone", "Address"])

                for user in self.data_manager.users.values():
                    profile = user.get_profile()
                    writer.writerow([
                        profile['id'], profile['full_name'], profile['email'],
                        profile['role'], profile['created_at'], profile['phone'], profile['address']
                    ])

            # Students CSV
            students_file = os.path.join(directory, "students.csv")
            with open(students_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["User ID", "Grade", "Subjects", "Average Grade"])

                for student in self.data_manager.students.values():
                    avg_grade = student.calculate_average_grade()
                    writer.writerow([
                        student._id, student.grade,
                        ", ".join(student.subjects.keys()), avg_grade
                    ])

            # Teachers CSV
            teachers_file = os.path.join(directory, "teachers.csv")
            with open(teachers_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["User ID", "Subjects", "Classes", "Workload"])

                for teacher in self.data_manager.teachers.values():
                    writer.writerow([
                        teacher._id, ", ".join(teacher.subjects),
                        ", ".join(teacher.classes), teacher.workload
                    ])

            # Assignments CSV
            assignments_file = os.path.join(directory, "assignments.csv")
            with open(assignments_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([
                    "ID", "Title", "Subject", "Teacher ID", "Class ID",
                    "Deadline", "Difficulty", "Submissions", "Grades"
                ])

                for assignment in self.data_manager.assignments.values():
                    writer.writerow([
                        assignment.id, assignment.title, assignment.subject,
                        assignment.teacher_id, assignment.class_id, assignment.deadline,
                        assignment.difficulty, len(assignment.submissions), len(assignment.grades)
                    ])

            self.log_export("CSV", directory, True)
            return True

        except Exception as e:
            logging.error(f"CSV export failed: {e}")
            self.log_export("CSV", directory, False)
            return False

    def export_to_sql(self, filename: str = "eduplatform_schema.sql") -> bool:
        """Export data as SQL scripts for SSMS"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                # Write database schema
                f.write("-- EduPlatform Database Schema for SQL Server\n")
                f.write("-- Generated on: " + datetime.now().isoformat() + "\n\n")

                # Create tables
                f.write("-- Create Users table\n")
                f.write("""CREATE TABLE Users
                           (
                               ID           INT PRIMARY KEY,
                               FullName     NVARCHAR(255) NOT NULL,
                               Email        NVARCHAR(255) UNIQUE NOT NULL,
                               PasswordHash NVARCHAR(255) NOT NULL,
                               Role         NVARCHAR(50) NOT NULL,
                               CreatedAt    DATETIME2 NOT NULL,
                               Phone        NVARCHAR(20),
                               Address      NVARCHAR(500)
                           );\n\n""")

                f.write("-- Create Students table\n")
                f.write("""CREATE TABLE Students
                           (
                               UserID       INT PRIMARY KEY,
                               Grade        NVARCHAR(10) NOT NULL,
                               Subjects     NVARCHAR(MAX),
                               AverageGrade FLOAT,
                               FOREIGN KEY (UserID) REFERENCES Users (ID)
                           );\n\n""")

                f.write("-- Create Teachers table\n")
                f.write("""CREATE TABLE Teachers
                           (
                               UserID   INT PRIMARY KEY,
                               Subjects NVARCHAR(MAX),
                               Classes  NVARCHAR(MAX),
                               Workload INT DEFAULT 0,
                               FOREIGN KEY (UserID) REFERENCES Users (ID)
                           );\n\n""")

                f.write("-- Create Assignments table\n")
                f.write("""CREATE TABLE Assignments
                           (
                               ID              INT PRIMARY KEY,
                               Title           NVARCHAR(255) NOT NULL,
                               Description     NVARCHAR(MAX),
                               Subject         NVARCHAR(100) NOT NULL,
                               TeacherID       INT       NOT NULL,
                               ClassID         NVARCHAR(50) NOT NULL,
                               Deadline        DATETIME2 NOT NULL,
                               Difficulty      NVARCHAR(20) DEFAULT 'medium',
                               SubmissionCount INT DEFAULT 0,
                               GradeCount      INT DEFAULT 0,
                               FOREIGN KEY (TeacherID) REFERENCES Users (ID)
                           );\n\n""")

                # Insert data
                f.write("-- Insert Users data\n")
                for user in self.data_manager.users.values():
                    profile = user.get_profile()
                    f.write(
                        f"INSERT INTO Users VALUES ({profile['id']}, N'{profile['full_name']}', '{profile['email']}', '{user._password_hash}', '{profile['role']}', '{profile['created_at']}', '{profile['phone']}', N'{profile['address']}');\n")

                f.write("\n-- Insert Students data\n")
                for student in self.data_manager.students.values():
                    avg_grade = student.calculate_average_grade()
                    subjects = ", ".join(student.subjects.keys())
                    f.write(
                        f"INSERT INTO Students VALUES ({student._id}, '{student.grade}', N'{subjects}', {avg_grade});\n")

                f.write("\n-- Insert Teachers data\n")
                for teacher in self.data_manager.teachers.values():
                    subjects = ", ".join(teacher.subjects)
                    classes = ", ".join(teacher.classes)
                    f.write(
                        f"INSERT INTO Teachers VALUES ({teacher._id}, N'{subjects}', N'{classes}', {teacher.workload});\n")

                f.write("\n-- Insert Assignments data\n")
                for assignment in self.data_manager.assignments.values():
                    f.write(
                        f"INSERT INTO Assignments VALUES ({assignment.id}, N'{assignment.title}', N'{assignment.description}', '{assignment.subject}', {assignment.teacher_id}, '{assignment.class_id}', '{assignment.deadline}', '{assignment.difficulty}', {len(assignment.submissions)}, {len(assignment.grades)});\n")

            self.log_export("SQL", filename, True)
            return True

        except Exception as e:
            logging.error(f"SQL export failed: {e}")
            self.log_export("SQL", filename, False)
            return False

    def auto_export_all(self) -> bool:
        """Export to all formats automatically"""
        xlsx_success = self.export_to_xlsx()
        csv_success = self.export_to_csv()
        sql_success = self.export_to_sql()

        return xlsx_success and csv_success and sql_success